import json
import openpyxl


class XlEquipmentParser:

    def __init__(self):
        self.rownum = 0
        self.db = []
        self.entities = []
        self.minds = []
        self.id = 1
        self.keys = set()

    def load(self, path):
        print(f"Open {path}")
        wb = openpyxl.load_workbook(path)
        for sheet in wb.worksheets:
            self.parse_sheet(sheet)
        print(self.db)
        wb.close()

    def parse_sheet(self, sheet):
        print(f"Opened sheet {sheet.title} with {sheet.max_row} rows and {sheet.max_column} columns")
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                v = sheet.cell(i, j).value
                if v is not None:
                    v = str(v).strip()
                    if (j, v) not in self.keys:
                        self.keys.add((j, v))
                        self.db.append((self.id, j - 1, v))
                        self.id += 1

    def convert(self):
        nblevel = max([item[1] for item in self.db]) + 1
        ids = []
        for i in range(nblevel):
            ids.append(-1)
        order = 0
        prevlevel = -1
        for item in self.db:
            id = item[0]
            level = item[1]
            short = item[2]
            pid = ids[level - 1] if level > 0 else -1
            label = short if level <= 2 else f"{[e for e in self.entities if e['id'] == pid][0]['label']}, {short}"
            ids[level] = id
            if level <= prevlevel != -1:
                self.entities[-1]["leaf"] = True
            entity = self.entity(id, label, short, pid, order)
            self.entities.append(entity)
            order += 1
            prevlevel = level
        self.entities[-1]["leaf"] = True

    def entity(self, id, label, shortlabel, parentid, order):
        return {
            "id": id,
            "label": label,
            "shortLabel": shortlabel,
            "parentId": parentid,
            "order": order,
            "leaf": False,
        }

    def save(self, path):
        js = json.dumps(self.entities, indent=True)
        print(js)
        with open(path, "w") as f:
            f.write(js)

    def convert_mind(self):
        nblevel = max([item[1] for item in self.db]) + 1
        ids = []
        for i in range(nblevel):
            ids.append(-1)
        even = True
        for item in self.db:
            id = item[0]
            level = item[1]
            short = item[2]
            pid = ids[level - 1] if level > 0 else -1
            ids[level] = id
            entity = self.entity_mind(id, short, pid)
            if pid == 0:
                if even:
                    entity["direction"] = "right"
                else:
                    entity["direction"] = "left"
                even = not even
            self.minds.append(entity)

    def entity_mind(self, id, topic, parentid):
        if parentid == -1:
            return {
                "id": "root",
                "isroot": True,
                "topic": topic,
            }
        else:
            return {
                "id": str(id),
                "topic": topic,
                "parentid": "root" if parentid == 0 else str(parentid),
            }

    def save_mind(self, path):
        dict = {"meta":
                {
                    "name": "demo",
                    "author": "contact@cyrilvincent.com",
                    "version": "0.1",
                },
                "format": "node_array",
                "data": self.minds}
        js = json.dumps(dict, indent=True)
        with open(path, "w") as f:
            f.write("var dbmind = ")
            f.write(js)


if __name__ == '__main__':
    print("XSLX2JSON")
    print("=========")
    path = "eq.xlsx"
    p = XlEquipmentParser()
    p.load(path)
    p.convert()
    path = path.replace("eq.xlsx", "db.json")
    p.save(path)
    p.convert_mind()
    path = path.replace("db.json", "mind.js")
    p.save_mind(path)
